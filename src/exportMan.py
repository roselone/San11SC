#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import *
from langconv import *
from conf import *
from kingdom import *
import sys
import random
import codecs
from ScenEditor import *
from functools import reduce

def readHero(fname):
    heroDic = {}
    f = codecs.open(fname, 'r', 'utf-8')
    heros = f.read().split('\n')
    for i in heros[:-1]:
        heroData = list(i.split(','))
        hid = int(heroData[0])
        heroDic[hid] = Hero(heroData)
    return heroDic

def findByName(name, heroDic):
    ret = -1
    name = name.strip()
    cname = Converter('zh-hant').convert(name)
    for _, hero in heroDic.items():
        if name == hero.name:
            ret = hero.hid
        elif name in special:
            ret = special[name]
        else:
            targetname = Converter('zh-hant').convert(hero.name)
            if cname == targetname:
                ret = hero.hid
        if ret != -1:
            break
    if ret == -1:
        print(name, ' give the id:')
        ret = int(input('>'))
    return ret
        
def findHerosID(playerID, heroNum, wb, kingID, heroDic):
    ws = wb.get_sheet_by_name('source')
    heroList = []
    for i in range(3, heroNum+3):
        heroID = findByName(ws.cell(row = i, column = playerID).value, heroDic)
        if heroID != kingID:
            heroList.append(heroID)
    heroList.sort()
    return heroList

def findCityID(cityName):
    ret = -1
    for cid, c in city.items():
        if cityName == c:
            ret = cid
            break
    if ret == -1:
        print(cityName, ' give the id:')
        ret = int(input('>'))
    return ret

def randomHero(heroDic, kingdoms):
    while True:
        heros = []
        for kingdom in kingdoms:
            heros.extend(kingdom.heroList)
            heros.append(kingdom.king)
            kingdom.tmp = []
        for i in range(670):
            if i not in heros:
                kingdoms[random.randint(0,7)].tmp.append(i)
        minh = 100
        maxh = 0
        for kingdom in kingdoms:
            minh = min(minh, len(kingdom.tmp))
            maxh = max(maxh, len(kingdom.tmp))
        if maxh - minh < 15:
            for kingdom in kingdoms:
                kingdom.heroList.extend(kingdom.tmp)
            break
    return kingdoms

def averageHero(heroDic, kingdoms):
    restHero = []
    tmpHero = []
    heros = []
    for kingdom in kingdoms:
        heros.extend(kingdom.heroList)
        heros.append(kingdom.king)
    for i in range(670):
        if i not in heros:
            if heroDic[i].halberdier == 'Ａ' or heroDic[i].halberdier == 'Ｓ':
                restHero.append(i)
            else:
                tmpHero.append(i)
    tmpHero = random.sample(tmpHero, len(tmpHero))
    restHero = random.sample(restHero, len(restHero))
    restHero.extend(tmpHero)
    cnt = 0
    for i in restHero:
        kingdoms[cnt].heroList.append(i)
        cnt = (cnt+1) % len(kingdoms)
    return kingdoms

def parseKingdom(playerNum, heroNum, wb, default):
    ws = wb.get_sheet_by_name('source')
    heroDic = readHero('hero.csv')
    kingdoms = []
    citys = []
    for i in range(1, playerNum+1):
        kingID = findByName(ws.cell(row = 1, column = i).value, heroDic)
        cid = findCityID(ws.cell(row = 2, column = i).value)
        citys.append(cid)
        kingdom = Kingdom(kingID,  cid, findHerosID(i, heroNum, wb, kingID, heroDic))
        kingdoms.append(kingdom)
    if default:
        kingdoms = randomHero(heroDic, kingdoms)
    else:
        kingdoms = averageHero(heroDic, kingdoms)
    return kingdoms

def main(playerNum, heroNum, inFile, default):
    wb = load_workbook(inFile)
    kingdoms = parseKingdom(playerNum, heroNum, wb, default)
    f = open('output.txt', 'w')
    for k in kingdoms:
        f.write(str(k))
        f.write('\n')
    f.close()
    genScen(kingdoms)
 
if __name__ == "__main__":
    default = 1
    if len(sys.argv) == 5 and sys.argv[4] == 'ave':
        default = 0
    main(int(sys.argv[1]), int(sys.argv[2]), sys.argv[3], default)
